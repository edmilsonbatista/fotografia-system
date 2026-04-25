"""
export_utils.py — Módulo de exportação PDF/Excel para Photo Pro Studio.
Funções puras que recebem dados e retornam BytesIO.
"""
from io import BytesIO
from datetime import date

import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm, mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT


# ---------------------------------------------------------------------------
# Formatação brasileira
# ---------------------------------------------------------------------------

def format_brl(value: float) -> str:
    """Formata float como moeda brasileira: R$ 1.234,56"""
    value = round(value, 2)
    formatted = f"{value:,.2f}"
    # Troca separadores para padrão BR
    formatted = formatted.replace(',', 'X').replace('.', ',').replace('X', '.')
    return f"R$ {formatted}"


def format_date_br(d) -> str:
    """Formata date como dd/mm/yyyy"""
    if d is None:
        return '-'
    try:
        return d.strftime('%d/%m/%Y')
    except (AttributeError, ValueError):
        return str(d)


MESES_PT = {
    1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
    5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
    9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
}


# ---------------------------------------------------------------------------
# PDF helpers
# ---------------------------------------------------------------------------

def _get_styles():
    """Retorna estilos padrão para os PDFs."""
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        'BrandTitle', parent=styles['Heading1'],
        fontSize=18, textColor=colors.HexColor('#333333'),
        alignment=TA_CENTER, spaceAfter=2 * mm
    ))
    styles.add(ParagraphStyle(
        'SubTitle', parent=styles['Normal'],
        fontSize=10, textColor=colors.HexColor('#666666'),
        alignment=TA_CENTER, spaceAfter=6 * mm
    ))
    styles.add(ParagraphStyle(
        'SectionTitle', parent=styles['Heading2'],
        fontSize=13, textColor=colors.HexColor('#333333'),
        spaceBefore=8 * mm, spaceAfter=4 * mm
    ))
    return styles


def _build_branding_elements(logo_path: str, subtitle: str = ''):
    """Retorna lista de flowables com logo + título da marca."""
    import os
    elements = []
    if logo_path and os.path.exists(logo_path):
        try:
            img = Image(logo_path, width=4 * cm, height=4 * cm)
            img.hAlign = 'CENTER'
            elements.append(img)
        except Exception:
            pass
    styles = _get_styles()
    elements.append(Paragraph('Photo Pro Studio', styles['BrandTitle']))
    if subtitle:
        elements.append(Paragraph(subtitle, styles['SubTitle']))
    elements.append(Spacer(1, 4 * mm))
    return elements


def _table_style_base():
    """Estilo base para tabelas PDF."""
    return TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4a90d9')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ])


# ---------------------------------------------------------------------------
# Excel — Eventos
# ---------------------------------------------------------------------------

def gerar_excel_eventos(eventos: list, data_atual: date) -> BytesIO:
    """Gera Excel (.xlsx) com a tabela de eventos e linha de totais."""
    colunas = ['Cliente', 'Tipo de Serviço', 'Data do Evento',
               'Ensaios Extras', 'Valor Negociado', 'Valor Pago', 'Status']

    rows = []
    for e in eventos:
        rows.append({
            'Cliente': e.get('cliente', ''),
            'Tipo de Serviço': e.get('tipo_servico', ''),
            'Data do Evento': format_date_br(e.get('data_evento')),
            'Ensaios Extras': e.get('ensaios_extras', 'Nenhum'),
            'Valor Negociado': e.get('valor_negociado', 0),
            'Valor Pago': e.get('valor_pago', 0),
            'Status': e.get('status', ''),
        })

    df = pd.DataFrame(rows, columns=colunas)

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Eventos')
        ws = writer.sheets['Eventos']

        # Larguras de coluna
        col_widths = [25, 20, 15, 18, 18, 18, 14]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w

        # Estilo do cabeçalho
        header_fill = PatternFill('solid', fgColor='4A90D9')
        header_font = Font(bold=True, color='FFFFFF', size=10)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Formatar colunas monetárias como BRL
        brl_font = Font(size=10)
        for row_idx in range(2, len(rows) + 2):
            for col_idx in [5, 6]:  # Valor Negociado, Valor Pago
                cell = ws.cell(row_idx, col_idx)
                cell.value = format_brl(cell.value if cell.value else 0)
                cell.font = brl_font
                cell.alignment = Alignment(horizontal='right')

        # Linha de totais
        total_row = len(rows) + 2
        total_neg = sum(e.get('valor_negociado', 0) for e in eventos)
        total_pago = sum(e.get('valor_pago', 0) for e in eventos)

        ws.cell(total_row, 1, 'TOTAIS').font = Font(bold=True, size=10)
        ws.cell(total_row, 5, format_brl(total_neg)).font = Font(bold=True, size=10)
        ws.cell(total_row, 5).alignment = Alignment(horizontal='right')
        ws.cell(total_row, 6, format_brl(total_pago)).font = Font(bold=True, size=10)
        ws.cell(total_row, 6).alignment = Alignment(horizontal='right')

        # Borda na linha de totais
        thin = Side(style='thin')
        for col_idx in range(1, 8):
            cell = ws.cell(total_row, col_idx)
            cell.border = Border(top=thin, bottom=thin)
            cell.fill = PatternFill('solid', fgColor='E8E8E8')

    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# PDF — Eventos
# ---------------------------------------------------------------------------

def gerar_pdf_eventos(eventos: list, data_atual: date, logo_path: str) -> BytesIO:
    """Gera PDF A4 landscape com tabela de eventos e branding."""
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5 * cm, rightMargin=1.5 * cm,
                            topMargin=1 * cm, bottomMargin=1 * cm)
    styles = _get_styles()
    elements = _build_branding_elements(
        logo_path,
        f'Relatório de Eventos — {format_date_br(data_atual)}'
    )

    if not eventos:
        elements.append(Spacer(1, 10 * mm))
        elements.append(Paragraph(
            'Nenhum evento encontrado para os filtros selecionados',
            ParagraphStyle('Empty', parent=styles['Normal'],
                           fontSize=12, alignment=TA_CENTER,
                           textColor=colors.HexColor('#888888'))
        ))
        doc.build(elements)
        buf.seek(0)
        return buf

    # Cabeçalho da tabela
    header = ['Cliente', 'Serviço', 'Data', 'Ensaios', 'Negociado', 'Pago', 'Status']
    data = [header]

    for e in eventos:
        data.append([
            e.get('cliente', ''),
            e.get('tipo_servico', ''),
            format_date_br(e.get('data_evento')),
            e.get('ensaios_extras', 'Nenhum'),
            format_brl(e.get('valor_negociado', 0)),
            format_brl(e.get('valor_pago', 0)),
            e.get('status', ''),
        ])

    # Linha de totais
    total_neg = sum(e.get('valor_negociado', 0) for e in eventos)
    total_pago = sum(e.get('valor_pago', 0) for e in eventos)
    data.append(['TOTAIS', '', '', '', format_brl(total_neg), format_brl(total_pago), ''])

    col_widths = [5.5 * cm, 3.5 * cm, 2.5 * cm, 3 * cm, 3.2 * cm, 3.2 * cm, 2.5 * cm]
    table = Table(data, colWidths=col_widths, repeatRows=1)

    style = _table_style_base()
    # Linha de totais em destaque
    last = len(data) - 1
    style.add('BACKGROUND', (0, last), (-1, last), colors.HexColor('#e8e8e8'))
    style.add('FONTNAME', (0, last), (-1, last), 'Helvetica-Bold')
    style.add('TEXTCOLOR', (0, last), (-1, last), colors.black)

    # Status colorido
    for i, e in enumerate(eventos, 1):
        status = e.get('status', '')
        if status == 'Realizado':
            style.add('TEXTCOLOR', (6, i), (6, i), colors.HexColor('#28a745'))
        elif status == 'Cancelado':
            style.add('TEXTCOLOR', (6, i), (6, i), colors.HexColor('#dc3545'))
        elif status == 'Agendado':
            style.add('TEXTCOLOR', (6, i), (6, i), colors.HexColor('#ffc107'))

    table.setStyle(style)
    elements.append(table)

    doc.build(elements)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Excel — Caixa
# ---------------------------------------------------------------------------

def gerar_excel_caixa(transacoes: list, pendentes: list,
                      entradas: float, saidas: float, saldo: float,
                      total_pendente: float, data_atual: date) -> BytesIO:
    """Gera Excel com aba Transações e aba Pendentes."""
    buf = BytesIO()

    # --- Aba principal: Transações ---
    colunas_tx = ['Data', 'Tipo', 'Descrição', 'Categoria', 'Valor']
    rows_tx = []
    for t in transacoes:
        rows_tx.append({
            'Data': format_date_br(t.get('data_transacao')),
            'Tipo': t.get('tipo', ''),
            'Descrição': t.get('descricao', ''),
            'Categoria': t.get('categoria', ''),
            'Valor': t.get('valor', 0),
        })
    df_tx = pd.DataFrame(rows_tx, columns=colunas_tx)

    # --- Aba Pendentes ---
    colunas_pend = ['Data Prevista', 'Cliente', 'Valor Pendente']
    rows_pend = []
    for p in pendentes:
        desc = p.get('descricao', '')
        cliente = desc.split(' - ')[1] if ' - ' in desc else desc
        rows_pend.append({
            'Data Prevista': format_date_br(p.get('data_transacao')),
            'Cliente': cliente,
            'Valor Pendente': p.get('valor', 0),
        })
    df_pend = pd.DataFrame(rows_pend, columns=colunas_pend)

    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df_tx.to_excel(writer, index=False, sheet_name='Transações')
        df_pend.to_excel(writer, index=False, sheet_name='Pendentes')

        # --- Estilizar aba Transações ---
        ws = writer.sheets['Transações']
        header_fill = PatternFill('solid', fgColor='4A90D9')
        header_font = Font(bold=True, color='FFFFFF', size=10)
        col_widths_tx = [15, 12, 35, 18, 18]
        for i, w in enumerate(col_widths_tx, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w
            ws[1][i - 1].fill = header_fill
            ws[1][i - 1].font = header_font
            ws[1][i - 1].alignment = Alignment(horizontal='center')

        # Formatar coluna Valor como BRL
        for row_idx in range(2, len(rows_tx) + 2):
            cell = ws.cell(row_idx, 5)
            cell.value = format_brl(cell.value if cell.value else 0)
            cell.alignment = Alignment(horizontal='right')

        # Linhas de resumo
        sr = len(rows_tx) + 3  # pula uma linha
        thin = Side(style='thin')
        summary_items = [
            ('Total de Entradas', entradas),
            ('Total de Saídas', saidas),
            ('Saldo Atual', saldo),
        ]
        for offset, (label, val) in enumerate(summary_items):
            r = sr + offset
            ws.cell(r, 4, label).font = Font(bold=True, size=10)
            ws.cell(r, 4).alignment = Alignment(horizontal='right')
            ws.cell(r, 5, format_brl(val)).font = Font(bold=True, size=10)
            ws.cell(r, 5).alignment = Alignment(horizontal='right')
            for c in [4, 5]:
                ws.cell(r, c).border = Border(top=thin, bottom=thin)
                ws.cell(r, c).fill = PatternFill('solid', fgColor='E8E8E8')

        # --- Estilizar aba Pendentes ---
        ws2 = writer.sheets['Pendentes']
        col_widths_pend = [15, 30, 18]
        for i, w in enumerate(col_widths_pend, 1):
            ws2.column_dimensions[ws2.cell(1, i).column_letter].width = w
            ws2[1][i - 1].fill = header_fill
            ws2[1][i - 1].font = header_font
            ws2[1][i - 1].alignment = Alignment(horizontal='center')

        for row_idx in range(2, len(rows_pend) + 2):
            cell = ws2.cell(row_idx, 3)
            cell.value = format_brl(cell.value if cell.value else 0)
            cell.alignment = Alignment(horizontal='right')

        # Total Pendente
        tr = len(rows_pend) + 3
        ws2.cell(tr, 2, 'Total Pendente').font = Font(bold=True, size=10)
        ws2.cell(tr, 2).alignment = Alignment(horizontal='right')
        ws2.cell(tr, 3, format_brl(total_pendente)).font = Font(bold=True, size=10)
        ws2.cell(tr, 3).alignment = Alignment(horizontal='right')
        for c in [2, 3]:
            ws2.cell(tr, c).border = Border(top=thin, bottom=thin)
            ws2.cell(tr, c).fill = PatternFill('solid', fgColor='E8E8E8')

    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# PDF — Caixa
# ---------------------------------------------------------------------------

def gerar_pdf_caixa(transacoes: list, pendentes: list,
                    entradas: float, saidas: float, saldo: float,
                    data_atual: date, logo_path: str) -> BytesIO:
    """Gera PDF do caixa com transações coloridas e seção de pendentes."""
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=1.5 * cm, rightMargin=1.5 * cm,
                            topMargin=1 * cm, bottomMargin=1 * cm)
    styles = _get_styles()
    elements = _build_branding_elements(
        logo_path,
        f'Controle de Caixa — {format_date_br(data_atual)}'
    )

    if not transacoes and not pendentes:
        elements.append(Spacer(1, 10 * mm))
        elements.append(Paragraph(
            'Nenhuma transação encontrada',
            ParagraphStyle('Empty', parent=styles['Normal'],
                           fontSize=12, alignment=TA_CENTER,
                           textColor=colors.HexColor('#888888'))
        ))
        doc.build(elements)
        buf.seek(0)
        return buf

    # --- Tabela de transações ---
    if transacoes:
        elements.append(Paragraph('Transações Realizadas', styles['SectionTitle']))
        header = ['Data', 'Tipo', 'Descrição', 'Categoria', 'Valor']
        data = [header]
        for t in transacoes:
            data.append([
                format_date_br(t.get('data_transacao')),
                t.get('tipo', ''),
                t.get('descricao', ''),
                t.get('categoria', ''),
                format_brl(t.get('valor', 0)),
            ])

        col_widths = [2.5 * cm, 2.2 * cm, 6 * cm, 3.5 * cm, 3 * cm]
        table = Table(data, colWidths=col_widths, repeatRows=1)
        style = _table_style_base()

        # Colorir tipo: verde para Entrada, vermelho para Saída
        for i, t in enumerate(transacoes, 1):
            tipo = t.get('tipo', '')
            if tipo == 'Entrada':
                style.add('TEXTCOLOR', (1, i), (1, i), colors.HexColor('#28a745'))
            elif tipo == 'Saída':
                style.add('TEXTCOLOR', (1, i), (1, i), colors.HexColor('#dc3545'))

        table.setStyle(style)
        elements.append(table)

        # Resumo
        elements.append(Spacer(1, 4 * mm))
        summary_data = [
            ['Total de Entradas', format_brl(entradas)],
            ['Total de Saídas', format_brl(saidas)],
            ['Saldo Atual', format_brl(saldo)],
        ]
        summary_table = Table(summary_data, colWidths=[10 * cm, 5 * cm])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f0f0f0')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(summary_table)

    # --- Seção Pendentes ---
    if pendentes:
        elements.append(Spacer(1, 6 * mm))
        elements.append(Paragraph('Entradas Pendentes', styles['SectionTitle']))
        header_pend = ['Data Prevista', 'Cliente', 'Valor Pendente']
        data_pend = [header_pend]
        for p in pendentes:
            desc = p.get('descricao', '')
            cliente = desc.split(' - ')[1] if ' - ' in desc else desc
            data_pend.append([
                format_date_br(p.get('data_transacao')),
                cliente,
                format_brl(p.get('valor', 0)),
            ])

        col_widths_pend = [3.5 * cm, 8 * cm, 3.5 * cm]
        table_pend = Table(data_pend, colWidths=col_widths_pend, repeatRows=1)
        style_pend = _table_style_base()
        table_pend.setStyle(style_pend)
        elements.append(table_pend)

    doc.build(elements)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# PDF — Relatório Mensal
# ---------------------------------------------------------------------------

def gerar_pdf_relatorio_mensal(mes: int, ano: int,
                                eventos: list, transacoes: list,
                                resumo: dict, logo_path: str) -> BytesIO:
    """Gera PDF do relatório financeiro mensal."""
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=1.5 * cm, rightMargin=1.5 * cm,
                            topMargin=1 * cm, bottomMargin=1 * cm)
    styles = _get_styles()
    nome_mes = MESES_PT.get(mes, str(mes))
    titulo = f'Relatório Financeiro Mensal — {nome_mes}/{ano}'
    elements = _build_branding_elements(logo_path, titulo)

    # --- Resumo financeiro ---
    elements.append(Paragraph('Resumo Financeiro', styles['SectionTitle']))
    summary_items = [
        ['Total de Entradas', format_brl(resumo.get('total_entradas', 0))],
        ['Total de Saídas', format_brl(resumo.get('total_saidas', 0))],
        ['Saldo do Mês', format_brl(resumo.get('saldo_mes', 0))],
        ['Total Negociado', format_brl(resumo.get('total_negociado', 0))],
        ['Total Recebido', format_brl(resumo.get('total_recebido', 0))],
        ['Total Pendente', format_brl(resumo.get('total_pendente', 0))],
    ]
    summary_table = Table(summary_items, colWidths=[10 * cm, 5 * cm])
    summary_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f0f0f0')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(summary_table)

    has_data = bool(eventos or transacoes)

    # --- Tabela de eventos ---
    elements.append(Paragraph('Eventos do Mês', styles['SectionTitle']))
    if eventos:
        header_ev = ['Cliente', 'Serviço', 'Data', 'Negociado', 'Pago', 'Status']
        data_ev = [header_ev]
        for e in eventos:
            data_ev.append([
                e.get('cliente', ''),
                e.get('tipo_servico', ''),
                format_date_br(e.get('data_evento')),
                format_brl(e.get('valor_negociado', 0)),
                format_brl(e.get('valor_pago', 0)),
                e.get('status', ''),
            ])
        col_w_ev = [4 * cm, 3 * cm, 2.5 * cm, 3 * cm, 3 * cm, 2.5 * cm]
        table_ev = Table(data_ev, colWidths=col_w_ev, repeatRows=1)
        table_ev.setStyle(_table_style_base())
        elements.append(table_ev)
    else:
        elements.append(Paragraph(
            'Nenhum registro encontrado para o período',
            ParagraphStyle('Empty', parent=styles['Normal'],
                           fontSize=10, alignment=TA_CENTER,
                           textColor=colors.HexColor('#888888'))
        ))

    # --- Tabela de transações ---
    elements.append(Paragraph('Transações do Mês', styles['SectionTitle']))
    if transacoes:
        header_tx = ['Data', 'Tipo', 'Descrição', 'Categoria', 'Valor']
        data_tx = [header_tx]
        for t in transacoes:
            data_tx.append([
                format_date_br(t.get('data_transacao')),
                t.get('tipo', ''),
                t.get('descricao', ''),
                t.get('categoria', ''),
                format_brl(t.get('valor', 0)),
            ])
        col_w_tx = [2.5 * cm, 2.2 * cm, 5.5 * cm, 3 * cm, 3 * cm]
        table_tx = Table(data_tx, colWidths=col_w_tx, repeatRows=1)
        style_tx = _table_style_base()
        for i, t in enumerate(transacoes, 1):
            tipo = t.get('tipo', '')
            if tipo == 'Entrada':
                style_tx.add('TEXTCOLOR', (1, i), (1, i), colors.HexColor('#28a745'))
            elif tipo == 'Saída':
                style_tx.add('TEXTCOLOR', (1, i), (1, i), colors.HexColor('#dc3545'))
        table_tx.setStyle(style_tx)
        elements.append(table_tx)
    else:
        if not eventos:
            pass  # Já mostrou a mensagem acima
        else:
            elements.append(Paragraph(
                'Nenhum registro encontrado para o período',
                ParagraphStyle('Empty2', parent=styles['Normal'],
                               fontSize=10, alignment=TA_CENTER,
                               textColor=colors.HexColor('#888888'))
            ))

    doc.build(elements)
    buf.seek(0)
    return buf
